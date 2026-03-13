"""
Скрипт для генерации Excel-шаблона импорта заявок.
Запустить один раз: python3 generate_template.py
После генерации файл можно удалить.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ============================================================
# Лист 1 — Шаблон импорта заявок
# ============================================================
ws = wb.active
ws.title = "Заявки"

# Стили
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
required_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
example_font = Font(italic=True, color="666666")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Заголовки колонок (только обязательные поля формы)
headers = [
    ("№", 5),
    ("Тип груза *", 35),
    ("Тип кузова ТС *", 25),
    ("Грузоподъемность ТС (тонн) *", 20),
    ("Стоимость перевозки (тенге) *", 22),
    ("Тариф перевозки *", 20),
    ("Дни отсрочки оплаты *", 18),
    ("Страхование груза *", 18),
    ("Дата/время погрузки *", 22),
    ("Адрес погрузки *", 45),
    ("Вес погрузки (тонн) *", 18),
    ("Объем погрузки (м3) *", 18),
    ("БИН грузоотправителя", 20),
    ("Контактное лицо (погрузка)", 25),
    ("Телефон (погрузка)", 18),
    ("Дата/время разгрузки *", 22),
    ("Адрес разгрузки *", 45),
    ("Вес разгрузки (тонн) *", 18),
    ("Объем разгрузки (м3) *", 18),
    ("БИН грузополучателя", 20),
    ("Контактное лицо (разгрузка)", 25),
    ("Телефон (разгрузка)", 18),
    ("Комментарий", 30),
]

# Записываем заголовки
for col_idx, (header, width) in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

ws.row_dimensions[1].height = 40

# ============================================================
# Data validation (выпадающие списки)
# ============================================================

# Тип груза (из БД, только is_active=true)
cargo_types = [
    "Автомашины, сельхозтехника и запчасти к ним",
    "Природные ресурсы (камень, руды и т. п.)",
    "Черные и цветные металлы",
    "Строительные материалы",
    "Алкоголь, табак",
    "Медицинское оборудование",
    "Компьютеры, оргтехника и компоненты к ним",
    "Грузы, требующие температурного режима",
    "Мебель",
    "Фармацевтическая продукция",
    "Полиграфические и канцелярские товары",
    "Продукты питания, не скоропортящиеся",
    "Оборудование для производства",
    "Бытовая химия/косметика",
    "Одежда, обувь",
    "Товары народного потребления (ТНП)",
    "Металлические изделия",
    "Бытовая техника",
    "Битум",
]

# Тип кузова ТС
body_types = [
    "Тентованный и шторный",
    "Рефрижераторный (-18)",
    "Изотермический",
    "Промтоварный фургон",
    "Автовоз",
    "Контейнеровоз",
    "Бортовой",
    "Цистерна",
    "Трал",
    "Рефрижератор (+2 +4)",
]

# Грузоподъемность (тонн)
capacity_values = [
    "0.7", "1", "1.5", "2", "3", "5", "7", "10", "20", "25", "26.5", "27"
]

# Тариф перевозки
tariff_types = [
    "Фрахт",
    "По времени работы",
    "По весу груза",
    "По объему груза",
]

# Дни отсрочки
payment_delays = ["0", "3", "7", "14", "21", "30", "45", "60", "90"]

# Страхование
insurance_options = ["Нет", "Да"]

# ============================================================
# Лист 2 — Справочники (для data validation)
# ============================================================
ws_ref = wb.create_sheet("Справочники")
ws_ref.sheet_state = 'hidden'  # Скрываем лист

def write_list(sheet, col, title, items):
    sheet.cell(row=1, column=col, value=title).font = Font(bold=True)
    for i, item in enumerate(items, 2):
        sheet.cell(row=i, column=col, value=item)
    return f"Справочники!${get_column_letter(col)}$2:${get_column_letter(col)}${len(items)+1}"

ref_cargo = write_list(ws_ref, 1, "Тип груза", cargo_types)
ref_body = write_list(ws_ref, 2, "Тип кузова", body_types)
ref_capacity = write_list(ws_ref, 3, "Грузоподъемность", capacity_values)
ref_tariff = write_list(ws_ref, 4, "Тариф", tariff_types)
ref_delay = write_list(ws_ref, 5, "Дни отсрочки", payment_delays)
ref_insurance = write_list(ws_ref, 6, "Страхование", insurance_options)

# Применяем data validation
dv_cargo = DataValidation(type="list", formula1=ref_cargo, allow_blank=False)
dv_cargo.error = "Выберите тип груза из списка"
dv_cargo.errorTitle = "Ошибка"
dv_cargo.prompt = "Выберите тип груза"
dv_cargo.promptTitle = "Тип груза"
ws.add_data_validation(dv_cargo)
dv_cargo.add(f"B2:B1000")

dv_body = DataValidation(type="list", formula1=ref_body, allow_blank=False)
dv_body.error = "Выберите тип кузова из списка"
ws.add_data_validation(dv_body)
dv_body.add(f"C2:C1000")

dv_capacity = DataValidation(type="list", formula1=ref_capacity, allow_blank=False)
dv_capacity.error = "Выберите грузоподъемность из списка"
ws.add_data_validation(dv_capacity)
dv_capacity.add(f"D2:D1000")

dv_tariff = DataValidation(type="list", formula1=ref_tariff, allow_blank=False)
dv_tariff.error = "Выберите тариф из списка"
ws.add_data_validation(dv_tariff)
dv_tariff.add(f"F2:F1000")

dv_delay = DataValidation(type="list", formula1=ref_delay, allow_blank=False)
dv_delay.error = "Выберите количество дней отсрочки"
ws.add_data_validation(dv_delay)
dv_delay.add(f"G2:G1000")

dv_insurance = DataValidation(type="list", formula1=ref_insurance, allow_blank=False)
dv_insurance.error = "Выберите Да или Нет"
ws.add_data_validation(dv_insurance)
dv_insurance.add(f"H2:H1000")

# ============================================================
# Примеры заполнения (2 строки)
# ============================================================

examples = [
    [
        1,
        "Битум",
        "Цистерна",
        "26.5",
        500000,
        "Фрахт",
        "30",
        "Нет",
        "2026-03-15 09:00",
        "г. Шымкент, ул. Толе би, промзона КазахБитум",
        26.5,
        30,
        "123456789012",
        "Иванов И.И.",
        "+77001234567",
        "2026-03-16 14:00",
        "Кыргызстан, Иссык-Кульская область, Джети-Огузский район, село Дархан",
        26.5,
        30,
        "987654321012",
        "Петров П.П.",
        "+77009876543",
        "",
    ],
    [
        2,
        "Строительные материалы",
        "Тентованный и шторный",
        "20",
        350000,
        "Фрахт",
        "14",
        "Да",
        "2026-03-17 08:00",
        "г. Алматы, ул. Абая 150, склад №3",
        18.5,
        25,
        "111222333444",
        "Сидоров С.С.",
        "+77051112233",
        "2026-03-18 16:00",
        "г. Караганда, промзона Майкудук, ул. Индустриальная 5",
        18.5,
        25,
        "555666777888",
        "Козлов К.К.",
        "+77054445566",
        "Хрупкий груз, требуется осторожная разгрузка",
    ],
]

for row_idx, example in enumerate(examples, 2):
    for col_idx, value in enumerate(example, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = example_font
        cell.fill = required_fill
        cell.border = thin_border
        if col_idx in (9, 16):  # Даты
            cell.number_format = 'YYYY-MM-DD HH:MM'
        elif col_idx in (5,):  # Стоимость
            cell.number_format = '#,##0'

# ============================================================
# Лист 3 — Инструкция
# ============================================================
ws_instr = wb.create_sheet("Инструкция")

instructions = [
    ("Инструкция по заполнению шаблона импорта заявок", True, 16),
    ("", False, 11),
    ("Общие правила:", True, 12),
    ("1. Каждая строка на листе «Заявки» = одна заявка на перевозку", False, 11),
    ("2. Поля, отмеченные * — обязательные для заполнения", False, 11),
    ("3. Поля с выпадающим списком — выбирайте значение из списка (не вводите вручную)", False, 11),
    ("4. Примеры (строки 2-3) можно удалить перед загрузкой", False, 11),
    ("", False, 11),
    ("Формат данных:", True, 12),
    ("Дата/время — формат: ГГГГ-ММ-ДД ЧЧ:ММ (пример: 2026-03-15 09:00)", False, 11),
    ("Вес — число в тоннах (пример: 26.5)", False, 11),
    ("Объем — число в кубометрах (пример: 30)", False, 11),
    ("Стоимость — число в тенге без пробелов (пример: 500000)", False, 11),
    ("БИН — 12 цифр (пример: 123456789012)", False, 11),
    ("Телефон — формат: +77XXXXXXXXX (пример: +77001234567)", False, 11),
    ("", False, 11),
    ("Описание полей:", True, 12),
    ("Тип груза * — категория перевозимого груза (выбрать из списка)", False, 11),
    ("Тип кузова ТС * — тип транспортного средства (выбрать из списка)", False, 11),
    ("Грузоподъемность ТС * — максимальная нагрузка в тоннах (выбрать из списка)", False, 11),
    ("Стоимость перевозки * — цена перевозки в тенге", False, 11),
    ("Тариф перевозки * — тип тарификации (выбрать из списка)", False, 11),
    ("Дни отсрочки оплаты * — количество дней отсрочки платежа (выбрать из списка)", False, 11),
    ("Страхование груза * — требуется ли страхование (Да/Нет)", False, 11),
    ("Адрес погрузки * — полный адрес точки погрузки", False, 11),
    ("Адрес разгрузки * — полный адрес точки разгрузки", False, 11),
    ("БИН грузоотправителя — БИН компании-отправителя (необязательно)", False, 11),
    ("БИН грузополучателя — БИН компании-получателя (необязательно)", False, 11),
    ("Комментарий — дополнительная информация о перевозке (необязательно)", False, 11),
    ("", False, 11),
    ("Допустимые значения:", True, 12),
    ("", False, 11),
    ("Тип груза:", True, 11),
]

for ct in cargo_types:
    instructions.append((f"  • {ct}", False, 11))

instructions.append(("", False, 11))
instructions.append(("Тип кузова ТС:", True, 11))
for bt in body_types:
    instructions.append((f"  • {bt}", False, 11))

instructions.append(("", False, 11))
instructions.append(("Грузоподъемность (тонн):", True, 11))
instructions.append((f"  {', '.join(capacity_values)}", False, 11))

instructions.append(("", False, 11))
instructions.append(("Тариф перевозки:", True, 11))
for tt in tariff_types:
    instructions.append((f"  • {tt}", False, 11))

instructions.append(("", False, 11))
instructions.append(("Дни отсрочки оплаты:", True, 11))
instructions.append((f"  {', '.join(payment_delays)}", False, 11))

for row_idx, (text, bold, size) in enumerate(instructions, 1):
    cell = ws_instr.cell(row=row_idx, column=1, value=text)
    cell.font = Font(bold=bold, size=size)

ws_instr.column_dimensions['A'].width = 80

# ============================================================
# Сохранение
# ============================================================
output_path = "/Users/admin/www/coube/coube-documentation/tasks/shablon_importa_zayavok.xlsx"
wb.save(output_path)
print(f"Шаблон создан: {output_path}")
